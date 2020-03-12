import csv
import openpyxl


def getexceldata(sheets):
    list1 = []
    # 循环遍历所有sheet
    for t in range(len(sheets)):
        sheet = wb.get_sheet_by_name(sheets[t])

        print('\n\n第' + str(t + 1) + '个sheet: ' + sheet.title + '->>>')

        # len_row代表表中有多少行数据，len_column代表excel表中有多少列数据
        len_row = sheet.max_row
        len_column = sheet.max_column
        # 合并的时候只保留第一张表的表头部分
        if t == 0:
            for i in range(2, len_row+1):
                list2 = []
                if sheet.cell(row=i, column=2).value is not None:
                    for j in [2, 3,7,11]:
                        list2.append(sheet.cell(row=i, column=j).value)
                    list1.append(list2)
        else:
            for i in range(2, len_row+1):
                list2 = []
                if sheet.cell(row=i, column=2).value is not None:
                    for j in [2, 3,7,11]:
                        list2.append(sheet.cell(row=i, column=j).value)
                    list1.append(list2)
    return list1


def dealwithlist(list1):
    currrowlength = list1.__len__()
    newlist = []
    for i in range(currrowlength):
        newlist.append([])
        for j in range(currrowlength):
            newlist[i].append(None)

    for row in range(0,7):
        newlist[row][0] = list1[row][0]

    currcolumn = 0
    for row in range(0, currrowlength, 7):
        if row == list1.__len__():
            break
        for j in range(1,4):
            for i in range(0, 7):
                newlist[i][j+currcolumn] = list1[row+i][j]
        currcolumn += 3
    return newlist


def transpose(matrix):
    new_matrix = []
    for i in range(len(matrix[0])):
        matrix1 = []
        for j in range(len(matrix)):
            matrix1.append(matrix[j][i])
        new_matrix.append(matrix1)
    return new_matrix


def saveascsv(name,res):
    with open(name, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(res)
        f.close()


if __name__ == '__main__':

    wb = openpyxl.load_workbook('test.xlsx')
    # 获取workbook中所有的表格
    sheets = wb.get_sheet_names()

    excellist = getexceldata(sheets)
    newlist = dealwithlist(excellist)
    res = transpose(newlist)
    saveascsv("test.csv", res)
